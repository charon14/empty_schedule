

import openpyxl
from pprint import pprint
from lxml import etree
from bs4 import BeautifulSoup
import re
import requests
import datetime
import os

file_read_path = "用户信息.xlsx"
users = []
schedule_info = []
week_num = 16

# 获取html文件
def askURL(user):
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,'
                  'application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'Content-Length': '219',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Host': 'jiaowu.sicau.edu.cn',
        'Origin': 'http://jiaowu.sicau.edu.cn',
        'Referer': 'http://jiaowu.sicau.edu.cn/web/web/web/index.asp',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36'
    }

    Requests = requests.session()

    a = Requests.get('http://jiaowu.sicau.edu.cn/web/web/web/index.asp')
    sign = etree.HTML(a.text).xpath('//input[@name="sign"]/@value')[0]
    hour_key = etree.HTML(a.text).xpath('//input[@name="hour_key"]/@value')[0]
    payload = 'user=%s&pwd=%s&lb=S&submit=&sign=%s&hour_key=%s' % (user[1], user[2], sign, hour_key)
    url = 'http://jiaowu.sicau.edu.cn/jiaoshi/bangong/check.asp'
    response = Requests.request("POST", url, headers=headers, data=payload)  # 登录教务网

    response.encoding = "utf-8"
    url = 'http://jiaowu.sicau.edu.cn/xuesheng/gongxuan/gongxuan/xuankeshow.asp?title_id1=3'
    a = Requests.get(url)
    content = a.content
    html = content.decode('gb2312', 'ignore')  # 得到课程表网页html

    return html


# 读取excel表中的信息
def read_excel(file_read_path):
    users = []
    user = []
    # 载入表
    wb = openpyxl.load_workbook(file_read_path)
    # 获取第一张表
    ws = wb.worksheets[0]
    # 从第二行开始遍历表
    for row in ws.iter_rows(2):
        for cell in row:
            user.append(str(cell.value))
        users.append(user)
        user = []
    # print(users)
    return users


# 获取单人的课表信息
def get_schedule_info(user):
    html = askURL(user)
    soup = BeautifulSoup(html, "html.parser")
    datalist = {}

    find_class_time_Info = re.compile(r'<td width="100">(.*?)</td>', re.S)
    find_week_info = re.compile(r'<td width="60">(.*?)</td>')

    try:
        # print(soup.prettify())
        table = soup.find_all('table', {'align': "center", 'border': "1", 'bordercolor': "#66CCCC", 'cellpadding': "2",
                                        'cellspacing': "0", 'style': "border-collapse: collapse", 'width': "1000"})[0]
        table = str(table)
        soup = BeautifulSoup(table, "html.parser")
    except:
        print("错误！请检查 %s 的账号密码是否正确"%user[0])
        print("程序已结束，请更正数据后重新打开程序")
        exit()

    for tr in soup.find_all('tr'):
        tr = str(tr)
        class_time_info = []
        week_info = []
        try:

            t1 = re.findall(find_class_time_Info, tr)[0]
            t1 = t1.replace("<br/>", ",")[:-1]
            class_time_info = t1.split(",")

            t2 = re.findall(find_week_info, tr)[1]
            if '-' in t2:
                start = int(t2.split('-')[0])
                end = int(t2.split('-')[1])
                for i in range(start, end + 1):
                    week_info.append(i)
            elif ',' in t2:
                week_info = list(map(int, t2.split(',')))
            else:
                week_info.append(int(t2))

        except:
            pass
            # print("error!")
        else:
            # print(class_time_info)
            # print(week_info)
            if class_time_info[0] != "":
                for v, i in enumerate(class_time_info):
                    if v < len(class_time_info) - 1:
                        if "(双)" in i or "(双)" in class_time_info[v + 1]:
                            if "(双)" in i:
                                i = i[:-3]
                            week_info = [j for j in week_info if j % 2 == 0]
                            if i not in datalist:
                                datalist[i] = week_info[:]
                            else:
                                datalist[i].extend(week_info[:])
                        elif "(单)" in i or "(单)" in class_time_info[v + 1]:
                            if "(单)" in i:
                                i = i[:-3]
                            week_info = [j for j in week_info if j % 2 == 1]
                            if i not in datalist:
                                datalist[i] = week_info[:]
                            else:
                                datalist[i].extend(week_info[:])
                        else:
                            if i not in datalist:
                                datalist[i] = week_info[:]
                            else:
                                datalist[i].extend(week_info[:])
                    else:
                        if "(双)" in i:
                            i = i[:-3]
                            week_info = [j for j in week_info if j % 2 == 0]
                            if i not in datalist:
                                datalist[i] = week_info[:]
                            else:
                                datalist[i].extend(week_info[:])
                        elif "(单)" in i:
                            i = i[:-3]
                            week_info = [j for j in week_info if j % 2 == 1]
                            if i not in datalist:
                                datalist[i] = week_info[:]
                            else:
                                datalist[i].extend(week_info[:])
                        else:
                            if i not in datalist:
                                datalist[i] = week_info[:]
                            else:
                                datalist[i].extend(week_info[:])

    # pprint(datalist)

    return datalist


# 处理成空课表
def schedule_processing(schedule_info):
    # pprint(schedule_info)
    schedule_list = []
    num_people = len(users)
    for i in range(1, 8):
        for j in range(1, 11):
            t1 = str(i) + '-' + str(j)
            # print(t1)
            schedule_list_single = []
            for v, p in enumerate(schedule_info):
                if t1 in p:
                    list1 = []
                    for k in range(1, week_num):
                        if k not in p[t1]:
                            list1.append(str(k))
                    if len(list1) !=0:
                        schedule_list_single.append(users[v][0] + '(' + ",".join(list1) + ')')
                else:
                    schedule_list_single.append(users[v][0])
            schedule_list.append(schedule_list_single)
    # print(schedule_list)

    return schedule_list


# 将空课表写入excel
def write_excel(empty_schedule):
    wb = openpyxl.Workbook()
    ws = wb.active

    for k in range(1,11):
        str1 = 'A'+str(k+1)
        str2 = '第'+str(k)+'节'
        ws[str1]=str2
    ws['B1']='星期一';ws['C1']='星期二';ws['D1']='星期三';ws['E1']='星期四';ws['F1']='星期五';ws['G1']='星期六';ws['H1']='星期天'

    i = 2
    j = 2
    for s in empty_schedule:
        s = chr(10).join(s)
        ws.cell(row=i, column=j, value=s)
        if i == 11:
            i = 2
            j += 1
        else:
            i+=1
    try:
        wb.save('空课表.xlsx')
    except:
        try:
            os.remove('空课表.xlsx')
            wb.save('空课表.xlsx')
        except PermissionError:
            print('空课表处于打开状态中，请关闭后再试')


def date2schedule(p2,schedule):
    p2.insert(0,2021)
    time_date = datetime.date(*p2).isocalendar()
    week = time_date[1] - 34
    day = time_date[2]
    t1 = 10 * (day - 1)
    t2 = 10 * day
    free_list = {}
    for t in range(t1, t2):
        # print(t)
        # print(schedule[t])
        s1 = str(t % 10 + 1)
        for d in schedule[t]:
            if '(' not in d:
                if d not in free_list:
                    free_list[d] = "第%s节 " % s1
                else:
                    free_list[d] += "第%s节 " % s1
            else:
                list1 = "".join(re.findall('[^\x00-\xff]', d))
                list2 = d.replace(list1, "")
                list2 = list2[1:-1]
                list2 = list(map(int, list2.split(",")))
                # print(list2)
                if week in list2:
                    if list1 not in free_list:
                        free_list[list1] = "第%s节 " % s1
                    else:
                        free_list[list1] += "第%s节 " % s1

    for key,value in free_list.items():
        print(key+":"+value)
    print()


if __name__ == "__main__":
    while 1:
        p1 = input("请输入A或者B来选择不同的模式,输入Q退出\nA:生成空课表\nB:输入日期查询当日没课的人\n")
        if p1 == "A" or p1 == 'a':
            users = read_excel(file_read_path)
            for user in users:
                schedule_info_single = get_schedule_info(user)
                schedule_info.append(schedule_info_single)
                print("已获取 %s 的课表信息"%user[0])
            # pprint(schedule_info)
            print("正在计算空课表中...")
            empty_schedule = schedule_processing(schedule_info)

            write_excel(empty_schedule)
            print("空课表表格已制作完成")
            print()
            users = []
            schedule_info = []
            with open('config.txt', 'w') as f:
                f.write(str(empty_schedule))

        if p1 == "B" or p1 == 'b':
            try:
                with open('config.txt', 'r') as f:
                    schedule = eval(f.read())
            except:
                print("没有对应的配置文件，请选择模式A生成配置文件后继续")
            else:
                try:
                    p2 = input("请输入日期，用英文逗号隔开（例如 9,10 ）\n")
                    p2 = list(map(int,p2.split(",")))
                    date2schedule(p2,schedule)
                except:
                    print("输入错误，请重新输入")
        if p1 == "Q" or p1 == 'q':
            exit()
        if p1 == 'week_num':
            week_num = int(p1)
